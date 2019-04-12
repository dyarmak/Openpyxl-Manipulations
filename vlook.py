import os
import openpyxl
from excelFNames import combinedFName
from myxlutils import get_column_names_and_index
from openpyxl.utils import get_column_letter


wbCombined = openpyxl.load_workbook(combinedFName)
sCombined = wbCombined.active
combinedDict = {}
get_column_names_and_index(sCombined, combinedDict)
column = get_column_letter(combinedDict["SubProjectTypeName"])

for r in range(2, sCombined.max_row+1):
        if(sCombined.cell(row=r, column = combinedDict["MasterProjectName"]).value == "Software"):
                sCombined.cell(row=r, column = combinedDict["Type"]).value = "Software" 

        elif(sCombined.cell(row=r, column = combinedDict["MasterProjectName"]).value == "HSE"):
                sCombined.cell(row=r, column = combinedDict["Type"]).value = "HSE" 

        elif(sCombined.cell(row=r, column = combinedDict["MasterProjectName"]).value == "CRB"):
                sCombined.cell(row=r, column = combinedDict["Type"]).value = "CRB" 

        else:
                vlookString = "=VLOOKUP(" + column + str(r) + ",vlook!A:B,2,FALSE)"
                sCombined.cell(row=r, column= combinedDict["Type"]).value = vlookString
        
wbCombined.save(combinedFName)
wbCombined.close()