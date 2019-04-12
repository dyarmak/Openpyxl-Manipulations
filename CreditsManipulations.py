import os
import glob
import datetime
import time
import xlrd
from openpyxl import Workbook
import openpyxl
from myxlutils import save_and_reopen, format_date_rows, get_column_names_and_index
from excelFNames import creditFName, logMe

startTimer = time.time()


# ------------- Load ---------------
wbCred = openpyxl.load_workbook(creditFName)
sCred = wbCred.active


# ----------- BEGIN Find Column Indexes -------------------
print("Loading column names")
creditsDictLog = open("creditsDictLog.txt", "w+")
# Should have a dict list that stores the column names
# maybe a key value pair with SubProjectStatus as the key, and the value as the column index

#Create an empty dictionary
creditsDict = {}
# fill the dictionary with:
# Key = column names
# Value = column index

get_column_names_and_index(sCred, creditsDict)
if(logMe == 1):
    #Prints all of the dictionary key:value pairs
    for field, row in creditsDict.items():
        logString = field + " : " + str(row) + "\n"
        creditsDictLog.write(logString)
        # print(logString)
    creditsDictLog.close()

    print("Created creditsDictLog.txt\n")
# ----------- END Find Column Indexes ------------------    



# Append "CR" to SubProjectID

for r in range(2, sCred.max_row+1):
    sCred.cell(row=r, column=creditsDict["SubProjectID"]).value = str(sCred.cell(row=r, column=creditsDict["SubProjectID"]).value) +"CR"
print("Added 'CR' after SubProjectIDs")

# ------------- INSERT Column ----------
#Insert after SubProjectTypeName
sCred.insert_cols(creditsDict["SubProjectTypeName"]+1)
sCred.cell(row=1, column= creditsDict["SubProjectTypeName"]+1).value = "Type"

# ------------ SAVE Output Excel File -------------

wbCred.save(creditFName)
print("File saved as " + creditFName)
wbCred.close()


# -------------- Timer -----------------------
endTimer = time.time()
creditsTime = endTimer-startTimer
print("Execution on Credits took: " + str(creditsTime) + "seconds to execute\n")
