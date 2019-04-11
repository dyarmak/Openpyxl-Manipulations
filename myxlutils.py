import openpyxl
import xlrd
from openpyxl import Workbook


def cvt_xls_to_xlsx(src_file_path, dst_file_path):
    """
    Takes .xls file as input and output .xlsx file with cell contents copied

    OpenPyXL can only work with .xlsx files (Excel 2007 and newer)

    """
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0,len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)


def get_column_names_and_index(sheetVar, emptyDict):
    """
    Fill an empty dictionary with the column names
    sheetVar = variable containing sheet
    emptyDict = an empty dictionary to hold column names
    """
    for x in range(1,(sheetVar.max_column+1)):
        emptyDict[str(sheetVar.cell(row=1, column=x).value)] = x
    
    print("Values added to dictionary")


def format_date_rows(sheetVar, colNamesDict, formatString, *args):
    """
    Apply date formatting to rows in dictionary
    sheetVar = variable containing sheet
    colNamesDict = dictionary of all column names in sheet
    formatString = Should be "mm-dd-yy" for it to work
    *args = the columnNames we want to apply date formatting to
    """
    
    for r in range(2, sheetVar.max_row+1):
        for columnName in args:
            sheetVar.cell(row=r, column=colNamesDict[columnName]).number_format = formatString

    print("date rows formatted")

def save_and_reopen(wbVar, sheetVar, tempFileName):
    """
    Save a temp file and reopen it to apply the date formatting
    wbVar = current woorkbook variable
    sheetVar = current sheet variable
    tempFileName = filename String
    """
    if(tempFileName.endswith('.xlsx')):
        wbVar.save(tempFileName)
        print(tempFileName + " was created")
        wbVar.close()
        wbVar = openpyxl.load_workbook(tempFileName)
        sheetVar = wbVar.active
        print("Closed and Re-Opened workbook so dates play nicely")
    else:
        print(tempFileName + " must be .xlsx")
        