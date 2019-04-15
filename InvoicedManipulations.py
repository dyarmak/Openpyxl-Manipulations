import os
import glob
import datetime
import time
import xlrd
from openpyxl import Workbook
import openpyxl
from myxlutils import save_and_reopen, format_date_rows, get_column_names_and_index
from excelFNames import invoicedFName, logMe

startTimer = time.time()

# ------------- Load ---------------
wbInvo = openpyxl.load_workbook(invoicedFName)
sInvo = wbInvo.active 

# ----------- BEGIN Find Column Indexes -------------------
# Should have a dictionary that stores the column names and index position

print("Loading column names")
if(logMe == 1):
    invoicedDictLog = open("invoicedDictLog.txt", "w+")

# Create an empty dictionary
invoicedDict = {}
# fill the dictionary with:
# Key = column names
# Value = column index

get_column_names_and_index(sInvo, invoicedDict)
if(logMe == 1):
    # Log all of the dictionary key:value pairs
    for field, row in invoicedDict.items():
        logString = field + " : " + str(row) + "\n"
        invoicedDictLog.write(logString)
        # print(logString)
    invoicedDictLog.close()
    print("Created invoicedDictLog.txt\n")
# ----------- END Find Column Indexes ------------------    

# ------------- Deferred Revenues ---------------

# IF InvoiceDateSent == 2018   Due Gets "Def-1"
if(logMe == 1):
    invoicedDeferredLog = open("invoicedDeferredLog.txt", "w+") 
for r in range(2, sInvo.max_row+1):
    if(sInvo.cell(row=r, column=invoicedDict["InvoiceDateSent"]).value.date().year == 2018):
        if(logMe == 1):
            logString = "SubProjectID " + str(sInvo.cell(row=r, column=invoicedDict["SubProjectID"]).value) + " marked as Deferred Revenue"
            invoicedDeferredLog.write(logString)
        sInvo.cell(row=r, column=invoicedDict["Due"]).value = "Def-1"
        

# ------------- INSERT Column ----------
#Insert after SubProjectTypeName
sInvo.insert_cols(invoicedDict["SubProjectTypeName"]+1)
sInvo.cell(row=1, column= invoicedDict["SubProjectTypeName"]+1).value = "Type"
# Update indexes after move
get_column_names_and_index(sInvo, invoicedDict)

# This is NOT in Christy's process, but in order to make the data actually line up we need to...
# Insert 3 rows after "OriginalDueDate"
for x in range(0,3):
    sInvo.insert_cols(invoicedDict["OriginalDueDate"]+1)
# Update indexes after move
get_column_names_and_index(sInvo, invoicedDict)

# ------------ SAVE Output Excel File -------------

wbInvo.save(invoicedFName)
print("File saved as " + invoicedFName)
wbInvo.close()

# -------------- Timer -----------------------
endTimer = time.time()
invoicedTime = endTimer-startTimer
print("Execution on Invoiced took: " + str(invoicedTime) + "seconds to execute\n")
