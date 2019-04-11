
# 1. read all the files in the directory
# 1a. convert any .xls files to .xlsx
# 2. load each .xlsx file into a variable
# 3. Perform Christy's manipulations
# Amalgamate the 3 queries into one?
# 4. Open the old ForecastYYYYMMDD.xlsx file
# 4a. rename to todays date
# 5. Delete the data in the detail sheet 
# 6. copy - paste contents from new queries  

from myxlutils import cvt_xls_to_xlsx, format_date_rows, get_column_names_and_index
import os
import openpyxl


# Make Output Folder
savePath = "py_Output"
if os.path.exists(savePath) is False:
        os.mkdir(savePath)
        
# OpenPyXL can only work with .xlsx files (Excel 2007 and newer)
# IF the files are of .xls type, we open them with xlrd and 
# copy the cell contents to a new .xlsx workbook
# save the files in the py_output folder
for filename in os.listdir('.'):
        if filename.endswith(".xls"):
                name, ext = filename.split('.')
                outName = name + ".xlsx"
                outPath = savePath + "\\" + outName
                cvt_xls_to_xlsx(filename, outPath)

# Do the .xls type (DATE) conversions here instead of in the files... 

# Open Forecast 
# Loop through all files in directory
for filename in os.listdir('.'):
    # only try to open those with .xlsx
    if filename.endswith(".xlsx"):
    # test filename then open book and sheet to according variables
        if filename.startswith("qry_Forecast"):
            wbFore = openpyxl.load_workbook(filename)
            sFore = wbFore.active
        #     print("qry_Forecast loaded")


# Load column names into a dict
forecastDict = {}
get_column_names_and_index(sFore, forecastDict)

# Apply forecast data column formatting
format_date_rows(sFore, forecastDict, "mm-dd-yy", "Due Date", "InvoiceDateSent", "OriginalDueDate")

# Save and close
wbFore.save("Forecast.xlsx")
wbFore.close()


# Open Credits
for filename in os.listdir('.'):
    # only try to open those with .xlsx
    if filename.endswith(".xlsx"):
    # test filename then open book and sheet to according variables
        if filename.startswith("qry_Credits"):
            wbCred = openpyxl.load_workbook(filename)
            sCred = wbCred.active
        #     print("qry_Credits loaded")

# Load column names into a dict
creditsDict = {}
get_column_names_and_index(sCred, creditsDict)

# Apply Credits date column formatting
format_date_rows(sCred, creditsDict, "mm-dd-yy", "Due Date", "OriginalDueDate")

# Save and close
wbCred.save("Credits.xlsx")
wbCred.close()


# Open Invoiced
# Loop through all files in directory
for filename in os.listdir('.'):
    # only try to open those with .xlsx
    if filename.endswith(".xlsx"):
    # test filename then open book and sheet to according variables
        if filename.startswith("qry_invoiced"):
            wbInvo = openpyxl.load_workbook(filename)
            sInvo = wbInvo.active 
        #     print("qry_invoiced loaded")

# Load column names into a dict
invoicedDict = {}
get_column_names_and_index(sInvo, invoicedDict)

# Apply Invoiced date column formatting
format_date_rows(sInvo, invoicedDict, "mm-dd-yy", "Due Date", "OriginalDueDate", "InvoiceDateSent")

# Save and close
wbInvo.save("Invoiced.xlsx")
wbInvo.close()

